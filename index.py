
import os
import sys
import cv2
import json
from openpyxl import Workbook
from openpyxl import load_workbook
from services.year_of_study import get_graduation_year
from services.filter import dsearch 

workbook = load_workbook(filename="Student_list.xlsx")
# sheet = workbook.active

images_dir = "source_images"
output_dir = "AWSImages"

crop_parameters = [
   [0,1150,0,2045,'front'],
   [0,1139,2102,2102+2026,'left'],
   [1183,1183+1139,0,2026,'right'],
]

list = os.listdir(images_dir)

student_list = []
pk = 2
for sheet in workbook.worksheets:
    for raw in sheet.iter_rows(min_row=2,min_col=1,values_only = True):
        print(raw)
        id,no,name,surname,gender,major,graduation_year = raw
        major = " ".join(major.split(" ")[2:])
        name = name.strip()
        if surname is None:
            surname = ""
        else:    
            surname = surname.strip()
        graduation_year = get_graduation_year(graduation_year)
        if major=="Communications & Media":
            major = "Communication & Media"
        student_list.append(
            {
                "model": "api.person",
                "pk": pk,
                "fields": {
                    "user": None,
                    "name": name.strip(),
                    "surname": surname.strip(),
                    "phone_number": None,
                    "profile_pic": None,
                    "on_campus": True,
                    "created_at": "2023-04-29T16:57:35.093Z",
                    "position": "Student",
                    "email": name.lower()+"."+surname.lower()+"_"+str(graduation_year)+"@ucentralasia.org",
                    "graduation_year": graduation_year,
                    "gender": gender,
                    "major": major,
                    "role":  None
                }
            }
        )
        pk= pk +1

aws_images = []
aws_pk = 1
for file_name in list:
    try:
        img = cv2.imread(images_dir+'/'+file_name)
        dir = file_name.split(".")[0]
        dirs = dir.split(" ")
        name,surname = dirs[0].strip(),dirs[1].strip()
        
        found = (dsearch(map(lambda x: {**x["fields"], "pk": x['pk']},student_list),name=name.strip(),surname=surname.strip()))
        if len(found) == 0:
            for student in student_list:
                full_name = student["fields"]["name"].strip() + " " + student["fields"]["surname"].strip()
                if full_name == dir.strip():
                    found = [student]
        if len(found) ==0:
            print("lost:", dir)

        if len(found) ==1:
            person = found[0]
            dir = str(person["pk"])
            try:
                os.mkdir(output_dir+"/"+ dir)
            except:
                pass
            for p in crop_parameters:
                cropped_image = img[p[0]:p[1],p[2]:p[3]]
                cv2.imwrite(output_dir+"/"+dir+"/"+p[4]+".jpg", cropped_image)
                aws_images.append({
                    "model": "api.awsimage",
                    "pk": aws_pk,
                    "fields": {
                        "person": person["pk"],
                        "image": "static/uploads/AWSImages/"+dir+"/"+p[4]+".jpg"
                    }
                })
                if p[4] == 'front':
                    cv2.imwrite("profile_pictures/_"+dir+"_generated_"+p[4]+".jpg", cropped_image)
                    student_list[person["pk"]-2]["fields"]["profile_pic"] = "static/uploads/profile_pictures/_"+dir+"_generated_"+p[4]+".jpg"
                aws_pk+=1
    except:
        print("Couldn't handle")
cv2.destroyAllWindows()
for i in student_list:
    if i["fields"]["profile_pic"] is None:
        print("Kosyachniki",i["fields"]["name"],i["fields"]["surname"])

with open("Person.json","w") as fp:
    json.dump(student_list,fp)

with open("AWSImage.json","w") as fp:
    json.dump(aws_images,fp)


